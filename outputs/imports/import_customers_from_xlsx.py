import json

import openpyxl

SOURCE_PATH = "/Applications/XAMPP/xamppfiles/htdocs/Phone_acc_imran/outputs/imports/customer.xlsx"
OUTPUT_PATH = "/Applications/XAMPP/xamppfiles/htdocs/Phone_acc_imran/outputs/imports/customers_import.json"

workbook = openpyxl.load_workbook(SOURCE_PATH, data_only=True)
worksheet = workbook.active

header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
headers = [str(value or "").strip().upper() for value in header_row]
header_map = {header: index for index, header in enumerate(headers)}

name_idx = header_map.get("SHOP NAME")
phone_idx = header_map.get("PHONE NUMBER")

if name_idx is None:
    raise SystemExit("SHOP NAME column not found in the first row header.")

items = []
for row in worksheet.iter_rows(min_row=2, values_only=True):
    name = row[name_idx] if name_idx < len(row) else None
    if name is None:
        continue
    name = str(name).strip()
    if not name:
        continue

    phone_value = None
    if phone_idx is not None and phone_idx < len(row):
        phone_value = row[phone_idx]

    phone = None
    if phone_value is not None:
        if isinstance(phone_value, (int, float)):
            phone = str(int(phone_value))
        else:
            phone = str(phone_value).strip()
            phone = phone or None

    items.append({
        "name": name,
        "phone": phone,
    })

with open(OUTPUT_PATH, "w", encoding="utf-8") as handle:
    json.dump(items, handle, ensure_ascii=True)

print(f"rows:{len(items)} file:{OUTPUT_PATH}")
