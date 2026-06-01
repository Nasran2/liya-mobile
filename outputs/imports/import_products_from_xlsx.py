import json

import openpyxl

SOURCE_PATH = "/Applications/XAMPP/xamppfiles/htdocs/Phone_acc_imran/outputs/imports/STROKE DETAILS.xlsx"
OUTPUT_PATH = "/Applications/XAMPP/xamppfiles/htdocs/Phone_acc_imran/outputs/imports/products_import.json"

workbook = openpyxl.load_workbook(SOURCE_PATH, data_only=True)
worksheet = workbook.active

header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
headers = [str(value or "").strip().upper() for value in header_row]
header_map = {header: index for index, header in enumerate(headers)}

product_idx = header_map.get("PRODUCT")
buying_idx = header_map.get("BUYING PRICE")
selling_idx = header_map.get("SALLING PRICE")
if selling_idx is None:
    selling_idx = header_map.get("SELLING PRICE")

if product_idx is None:
    raise SystemExit("PRODUCT column not found in the first row header.")

def to_number(value: object) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    return float(text) if text else 0.0

items = []
for row in worksheet.iter_rows(min_row=2, values_only=True):
    name = row[product_idx] if product_idx < len(row) else None
    if name is None:
        continue
    name = str(name).strip()
    if not name:
        continue

    cost = 0.0
    if buying_idx is not None and buying_idx < len(row):
        cost = to_number(row[buying_idx])

    selling = 0.0
    if selling_idx is not None and selling_idx < len(row):
        selling = to_number(row[selling_idx])

    items.append({
        "name": name,
        "cost_price": cost,
        "selling_price": selling,
    })

with open(OUTPUT_PATH, "w", encoding="utf-8") as handle:
    json.dump(items, handle, ensure_ascii=True)

print(f"rows:{len(items)} file:{OUTPUT_PATH}")
